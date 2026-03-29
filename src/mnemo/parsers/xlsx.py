"""XLSX parser — extracts sheet data and metadata via openpyxl."""

from __future__ import annotations

import hashlib
from pathlib import Path

from mnemo.parser import NoteDocument

try:
    from openpyxl import load_workbook as _load_workbook  # type: ignore[import-untyped]

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def parse_xlsx(file_path: Path, vault_root: Path | None = None) -> NoteDocument:
    """Parse an XLSX file into a NoteDocument.

    Requires the ``openpyxl`` library (``pip install openpyxl>=3.1``).
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for XLSX parsing. Install with: pip install 'mnemo-secondbrain[documents]'"
        )

    wb = _load_workbook(str(file_path), read_only=True, data_only=True)

    # --- extract text per sheet -----------------------------------------------
    sheets_text: list[str] = []
    headings: list[str] = []

    for sheet_name in wb.sheetnames:
        headings.append(sheet_name)
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            line = " | ".join(c for c in cells if c)
            if line:
                rows_text.append(line)
        if rows_text:
            sheets_text.append(f"## {sheet_name}\n" + "\n".join(rows_text))

    wb.close()

    body = "\n\n".join(sheets_text)

    # --- metadata / synthetic frontmatter ------------------------------------
    props = wb.properties if hasattr(wb, "properties") else None
    title = file_path.stem
    author = ""
    created = None
    if props:
        title = (getattr(props, "title", "") or "").strip() or file_path.stem
        author = (getattr(props, "creator", "") or "").strip()
        raw_created = getattr(props, "created", None)
        if raw_created:
            created = raw_created.isoformat() if hasattr(raw_created, "isoformat") else str(raw_created)

    frontmatter: dict = {
        "title": title,
        "type": "source",
        "format": "xlsx",
        "tags": ["xlsx", "external"],
        "sheets": wb.sheetnames,
    }
    if author:
        frontmatter["author"] = author
    if created:
        frontmatter["created"] = created

    # --- key ------------------------------------------------------------------
    if vault_root is not None:
        try:
            rel = file_path.resolve().relative_to(vault_root.resolve())
            key = rel.with_suffix("").as_posix()
        except ValueError:
            key = file_path.stem
    else:
        key = file_path.stem

    content = body
    checksum = hashlib.sha256(content.encode("utf-8")).hexdigest()[:16]

    return NoteDocument(
        path=file_path,
        name=file_path.stem,
        key=key,
        content=content,
        body=body,
        frontmatter=frontmatter,
        wiki_links=[],
        tags=frontmatter["tags"],
        headings=headings[:30],
        checksum=checksum,
    )
